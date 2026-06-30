"""
Importe l'inventaire matériel VPMI (Excel .xlsx ou CSV) dans la Liste SharePoint.

Conçu pour le format réel de VPMI :
  - deux sections : "MATERIEL ATTRIBUER" et "MATERIEL NON ATTRIBUE (EN STOCK)"
    -> renseigne automatiquement la colonne Statut (Attribué / En stock) ;
  - la catégorie (ORDINATEUR, CLE WIFI...) est en colonne A "au fil de l'eau"
    -> reportée sur les lignes suivantes ;
  - en-têtes parfois abîmés (ex. "MARQUES+B6:M7") -> reconnaissance souple ;
  - lignes "TOTAL" et lignes vides ignorées ;
  - extraction du N° de série quand il est noté dans le modèle ("SN ...").

Usage :
    python scripts/import_excel.py "chemin/fichier.xlsx"
    python scripts/import_excel.py "fichier.csv" --site "Côte d'Ivoire"
    python scripts/import_excel.py "fichier.xlsx" --sheet "Feuil1" --dry-run

Lancez d'abord scripts/create_list.py.
"""
import argparse
import csv
import re
import sys
import unicodedata
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# Console Windows : forcer UTF-8 pour les accents/emojis
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

from app.graph import GraphError, client  # noqa: E402


# ---------- normalisation ----------
def norm(s) -> str:
    s = ("" if s is None else str(s)).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


def resolve_header(h: str):
    """Associe un en-tête (même abîmé) à un nom de colonne interne, ou None."""
    n = norm(h)
    if not n:
        return None
    if n.startswith("marque"):
        return "Marque"
    if "model" in n or "modele" in n:
        return "Modele"
    if "quantit" in n:
        return "Quantite"
    if "etat" in n or "condition" in n:
        return "Etat"
    if "accessoir" in n:
        return "Accessoires"
    if "utilisateur" in n or "affecte" in n:
        return "Utilisateur"
    if "pays" in n or n.startswith("site") or "localisation" in n or "agence" in n:
        return "Site"
    if "remarque" in n or "commentaire" in n or "note" in n:
        return "Remarque"
    if n.startswith("categ") or n.startswith("type"):
        return "Categorie"
    return None


CATEGORY_MAP = {
    "ordinateur": "Ordinateur", "pc": "Ordinateur", "laptop": "Ordinateur",
    "cle wifi": "Clé WiFi", "cle wi-fi": "Clé WiFi", "cle usb": "Clé WiFi",
    "carte memoire": "Carte mémoire", "ecran": "Écran", "imprimante": "Imprimante",
    "telephone": "Téléphone", "serveur": "Serveur", "onduleur": "Onduleur",
}

ETAT_MAP = {
    "en cours d'utilisation": "En cours d'utilisation",
    "en cours dutilisation": "En cours d'utilisation",
    "fonctionne": "Fonctionne",
    "en reparation": "En réparation",
    "endommage": "Endommagé", "endommager": "Endommagé",
    "perdu": "Perdu",
    "hors service": "Hors service", "hs": "Hors service",
}

SITE_MAP = {
    "cote d'ivoire": "Côte d'Ivoire", "cote divoire": "Côte d'Ivoire",
    "cote d ivoire": "Côte d'Ivoire", "ci": "Côte d'Ivoire", "rci": "Côte d'Ivoire",
    "senegal": "Sénégal", "cameroun": "Cameroun", "cameroon": "Cameroun",
    "guadeloupe": "Guadeloupe", "martinique": "Martinique", "guyane": "Guyane",
    "siege": "Siège",
}

TOTAL_RE = re.compile(r"^\s*total", re.I)
SERIAL_RE = re.compile(r"\b(?:s\.?/?n|num[ée]ro de s[ée]rie)\b[:\s]*([A-Za-z0-9\-]+)", re.I)


def map_category(raw: str) -> str:
    return CATEGORY_MAP.get(norm(raw), str(raw).strip().capitalize())


def map_site(raw: str) -> str:
    return SITE_MAP.get(norm(raw), str(raw).strip())


def map_etat(raw: str) -> str:
    return ETAT_MAP.get(norm(raw), str(raw).strip())


def split_serial(modele: str):
    """Sépare un éventuel N° de série noté dans le modèle. Retourne (modele, serie|None)."""
    if not modele:
        return modele, None
    flat = " ".join(str(modele).split())
    m = SERIAL_RE.search(flat)
    if m:
        serie = m.group(1)
        cleaned = flat[: m.start()].strip(" -:")
        return (cleaned or flat), serie
    return flat, None


# ---------- lecture du fichier ----------
def read_rows(path: Path, sheet: str | None):
    if path.suffix.lower() == ".csv":
        # délimiteur ';' (export Excel FR) avec repli sur ','
        # newline="" : laisse csv gérer les cellules multi-lignes entre guillemets
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as fh:
            sample = fh.read(2000)
            fh.seek(0)
            delim = ";" if sample.count(";") >= sample.count(",") else ","
            return [list(r) for r in csv.reader(fh, delimiter=delim)]
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    print(f"   feuille : {ws.title}")
    return [list(r) for r in ws.iter_rows(values_only=True)]


def cell(row, i):
    return row[i] if i < len(row) and row[i] is not None else ""


# ---------- traitement ----------
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("file", help="Chemin du fichier (.xlsx ou .csv)")
    ap.add_argument("--sheet", help="Nom de la feuille (défaut : la première)")
    ap.add_argument("--site", help="Site/Pays à appliquer à toutes les lignes (ex: \"Côte d'Ivoire\")")
    ap.add_argument("--dry-run", action="store_true", help="Aperçu sans écrire dans SharePoint")
    args = ap.parse_args()

    path = Path(args.file)
    if not path.exists():
        print("❌ Fichier introuvable :", path)
        sys.exit(1)

    print(f"→ Lecture de '{path.name}'")
    rows = read_rows(path, args.sheet)

    current_statut = ""
    current_cat = ""
    current_site = args.site or ""   # valeur par défaut, écrasée par une colonne PAYS/Site
    header_map: dict[int, str] = {}
    parsed: list[dict] = []

    for row in rows:
        joined = norm(" ".join(str(c) for c in row if c))
        if not joined:
            continue

        # 1) marqueur de section -> statut
        if "non attribu" in joined or "en stock" in joined:
            current_statut = "En stock"
            header_map = {}
            continue
        if "attribu" in joined:  # "MATERIEL ATTRIBUER"
            current_statut = "Attribué"
            header_map = {}
            continue

        # 2) ligne de total -> ignorée
        if TOTAL_RE.match(str(cell(row, 0))) or joined.startswith("total"):
            continue

        # 3) ligne d'en-tête (contient marque + modele)
        targets = {i: resolve_header(c) for i, c in enumerate(row)}
        named = {i: t for i, t in targets.items() if t}
        if "Marque" in named.values() and ("Modele" in named.values()):
            header_map = {}
            for i, t in named.items():
                if t and t not in header_map.values():  # 1ère occurrence gagne
                    header_map[i] = t
            # une catégorie peut figurer en colonne A de l'en-tête
            c0 = cell(row, 0)
            if c0 and resolve_header(c0) is None:
                current_cat = map_category(c0)
            continue

        if not header_map:
            continue  # pas encore de structure connue

        # 4) catégorie au fil de l'eau (colonne A)
        c0 = cell(row, 0)
        if c0 and resolve_header(c0) is None:
            current_cat = map_category(c0)

        # 5) ligne de données
        rec: dict = {}
        for i, t in header_map.items():
            rec[t] = cell(row, i)

        marque = str(rec.get("Marque", "")).strip()
        modele = str(rec.get("Modele", "")).strip()
        if not marque and not modele:
            continue  # ligne sans matériel réel

        modele, serie = split_serial(modele)

        fields: dict = {"Categorie": current_cat or "Autre"}
        if marque:
            fields["Marque"] = marque
        if modele:
            fields["Modele"] = modele
        if serie:
            fields["NumeroSerie"] = serie
        if rec.get("Quantite") not in (None, ""):
            try:
                fields["Quantite"] = float(rec["Quantite"])
            except (TypeError, ValueError):
                pass
        if str(rec.get("Etat", "")).strip():
            fields["Etat"] = map_etat(rec["Etat"])
        if str(rec.get("Accessoires", "")).strip():
            fields["Accessoires"] = str(rec["Accessoires"]).strip()
        if str(rec.get("Utilisateur", "")).strip():
            fields["Utilisateur"] = str(rec["Utilisateur"]).strip()
        if str(rec.get("Remarque", "")).strip():
            fields["Remarque"] = " ".join(str(rec["Remarque"]).split())
        if current_statut:
            fields["Statut"] = current_statut
        # Pays/Site : colonne dédiée au fil de l'eau, sinon valeur de --site
        if str(rec.get("Site", "")).strip():
            current_site = map_site(rec["Site"])
        if current_site:
            fields["Site"] = current_site

        # Marque obligatoire dans le schéma : repli si vide
        fields.setdefault("Marque", "N/C")
        parsed.append(fields)

    if not parsed:
        print("❌ Aucune ligne de matériel reconnue. Vérifiez le fichier / la feuille.")
        sys.exit(1)

    print(f"   {len(parsed)} matériel(s) détecté(s).")
    if args.dry_run:
        for f in parsed:
            print("   [dry-run]", {k: v for k, v in f.items() if v})
        print(f"\nℹ️  Aperçu uniquement (--dry-run). Rien n'a été écrit.")
        return

    created, errors = 0, 0
    for f in parsed:
        try:
            client.create_item(f)
            created += 1
            if created % 20 == 0:
                print(f"   … {created} importés")
        except GraphError as e:
            errors += 1
            print("   ⚠️", e)

    print(f"\n✅ {created} matériel(s) importé(s), {errors} erreur(s).")


if __name__ == "__main__":
    try:
        main()
    except GraphError as e:
        print("❌", e)
        sys.exit(1)
