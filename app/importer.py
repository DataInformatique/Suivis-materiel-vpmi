"""
Analyse d'un inventaire matériel VPMI (lignes de tableur) → liste de champs.

Réutilisé par :
  - scripts/import_excel.py (ligne de commande),
  - l'endpoint /api/import (import depuis l'interface web).

Gère le format réel : sections ATTRIBUER / EN STOCK, catégorie au fil de l'eau,
colonne PAYS au fil de l'eau, en-têtes abîmés, lignes TOTAL, extraction du N° de série.
"""
from __future__ import annotations

import csv
import io
import re
import unicodedata
from typing import Any


# ---------- normalisation ----------
def norm(s: Any) -> str:
    s = ("" if s is None else str(s)).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


def resolve_header(h: Any):
    """Associe un en-tête (même abîmé) à un nom de colonne interne, ou None."""
    n = norm(h)
    if not n:
        return None
    if n.startswith("marque"):
        return "Marque"
    if "model" in n or "modele" in n:
        return "Modele"
    if "quantit" in n:
        return "Quantite"
    if "etat" in n or "condition" in n:
        return "Etat"
    if "accessoir" in n:
        return "Accessoires"
    if "utilisateur" in n or "affecte" in n:
        return "Utilisateur"
    if "pays" in n or n.startswith("site") or "localisation" in n or "agence" in n:
        return "Site"
    if "remarque" in n or "commentaire" in n or "note" in n:
        return "Remarque"
    if n.startswith("categ") or n.startswith("type"):
        return "Categorie"
    return None


CATEGORY_MAP = {
    "ordinateur": "Ordinateur", "pc": "Ordinateur", "laptop": "Ordinateur",
    "tablette": "Tablette", "tab": "Tablette",
    "telephone": "Téléphone", "phone": "Téléphone",
    "cle wifi": "Clé WiFi", "cle wi-fi": "Clé WiFi", "cle usb": "Clé WiFi",
    "carte memoire": "Carte mémoire", "ecran": "Écran", "imprimante": "Imprimante",
    "serveur": "Serveur", "onduleur": "Onduleur",
}

ETAT_MAP = {
    "en cours d'utilisation": "En cours d'utilisation",
    "en cours dutilisation": "En cours d'utilisation",
    "fonctionne": "Fonctionne",
    "en reparation": "En réparation",
    "endommage": "Endommagé", "endommager": "Endommagé",
    "perdu": "Perdu",
    "hors service": "Hors service", "hs": "Hors service",
}

SITE_MAP = {
    "cote d'ivoire": "Côte d'Ivoire", "cote divoire": "Côte d'Ivoire",
    "cote d ivoire": "Côte d'Ivoire", "ci": "Côte d'Ivoire", "rci": "Côte d'Ivoire",
    "senegal": "Sénégal", "cameroun": "Cameroun", "cameroon": "Cameroun",
    "guadeloupe": "Guadeloupe", "martinique": "Martinique", "guyane": "Guyane",
    "chartres": "Chartres", "siege": "Chartres", "france": "Chartres",
}

TOTAL_RE = re.compile(r"^\s*total", re.I)
SERIAL_RE = re.compile(r"\b(?:s\.?/?n|num[ée]ro de s[ée]rie)\b[:\s]*([A-Za-z0-9\-]+)", re.I)


def map_category(raw: Any) -> str:
    return CATEGORY_MAP.get(norm(raw), str(raw).strip().capitalize())


def map_site(raw: Any) -> str:
    return SITE_MAP.get(norm(raw), str(raw).strip())


def map_etat(raw: Any) -> str:
    return ETAT_MAP.get(norm(raw), str(raw).strip())


def split_serial(modele: str):
    """Sépare un éventuel N° de série noté dans le modèle. Retourne (modele, serie|None)."""
    if not modele:
        return modele, None
    flat = " ".join(str(modele).split())
    m = SERIAL_RE.search(flat)
    if m:
        serie = m.group(1)
        cleaned = flat[: m.start()].strip(" -:")
        return (cleaned or flat), serie
    return flat, None


def cell(row, i):
    return row[i] if i < len(row) and row[i] is not None else ""


def parse_rows(rows: list[list], default_site: str | None = None) -> list[dict]:
    """Transforme les lignes brutes d'un tableur en liste de matériels (champs internes)."""
    current_statut = ""
    current_cat = ""
    current_site = default_site or ""
    header_map: dict[int, str] = {}
    parsed: list[dict] = []

    for row in rows:
        joined = norm(" ".join(str(c) for c in row if c))
        if not joined:
            continue

        if "non attribu" in joined or "en stock" in joined:
            current_statut = "En stock"
            header_map = {}
            continue
        if "attribu" in joined:
            current_statut = "Attribué"
            header_map = {}
            continue

        if TOTAL_RE.match(str(cell(row, 0))) or joined.startswith("total"):
            continue

        named = {i: resolve_header(c) for i, c in enumerate(row)}
        named = {i: t for i, t in named.items() if t}
        if "Marque" in named.values() and "Modele" in named.values():
            header_map = {}
            for i, t in named.items():
                if t and t not in header_map.values():
                    header_map[i] = t
            c0 = cell(row, 0)
            if c0 and resolve_header(c0) is None:
                current_cat = map_category(c0)
            continue

        if not header_map:
            continue

        c0 = cell(row, 0)
        if c0 and resolve_header(c0) is None:
            current_cat = map_category(c0)

        rec: dict = {}
        for i, t in header_map.items():
            rec[t] = cell(row, i)

        marque = str(rec.get("Marque", "")).strip()
        modele = str(rec.get("Modele", "")).strip()
        if not marque and not modele:
            continue

        modele, serie = split_serial(modele)

        fields: dict = {"Categorie": current_cat or "Autre"}
        if marque:
            fields["Marque"] = marque
        if modele:
            fields["Modele"] = modele
        if serie:
            fields["NumeroSerie"] = serie
        if rec.get("Quantite") not in (None, ""):
            try:
                fields["Quantite"] = float(rec["Quantite"])
            except (TypeError, ValueError):
                pass
        if str(rec.get("Etat", "")).strip():
            fields["Etat"] = map_etat(rec["Etat"])
        if str(rec.get("Accessoires", "")).strip():
            fields["Accessoires"] = str(rec["Accessoires"]).strip()
        if str(rec.get("Utilisateur", "")).strip():
            fields["Utilisateur"] = str(rec["Utilisateur"]).strip()
        if str(rec.get("Remarque", "")).strip():
            fields["Remarque"] = " ".join(str(rec["Remarque"]).split())
        if current_statut:
            fields["Statut"] = current_statut
        if str(rec.get("Site", "")).strip():
            current_site = map_site(rec["Site"])
        if current_site:
            fields["Site"] = current_site

        fields.setdefault("Marque", "N/C")
        parsed.append(fields)

    return parsed


# ---------- lecture de fichiers (chemin ou contenu binaire) ----------
def rows_from_bytes(content: bytes, filename: str, sheet: str | None = None) -> list[list]:
    """Lit un .xlsx ou .csv depuis son contenu binaire → lignes brutes."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        delim = ";" if text[:2000].count(";") >= text[:2000].count(",") else ","
        return [list(r) for r in csv.reader(io.StringIO(text), delimiter=delim)]
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[sheet] if sheet else wb.active
    return [list(r) for r in ws.iter_rows(values_only=True)]
